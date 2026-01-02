import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter

def create_theme_summary(df):
    """
    주식 분석 데이터프레임에서 테마별 빈도수를 집계하여 요약 데이터프레임을 생성합니다.
    
    Args:
        df (pd.DataFrame): '테마_x' 컬럼들이 포함된 원본 데이터프레임
        
    Returns:
        pd.DataFrame: '테마', '종목갯수' 컬럼을 가진 요약 데이터프레임 (종목갯수 내림차순 정렬)
    """
    # 1. '테마_'로 시작하는 모든 컬럼 찾기
    theme_cols = [col for col in df.columns if col.startswith('테마_')]
    
    # 2. 데이터프레임을 '긴 형식(Long Format)'으로 변환 (melt)
    #    여러 열에 퍼져 있는 테마들을 하나의 '테마' 열로 모읍니다.
    melted_df = df.melt(value_vars=theme_cols, value_name='테마')
    
    # 3. 빈 값(NaN, 공백) 제거
    melted_df = melted_df.dropna(subset=['테마'])
    melted_df = melted_df[melted_df['테마'] != '']
    
    # 4. 테마별 종목 개수 세기
    summary_df = melted_df.groupby('테마').size().reset_index(name='종목갯수')
    
    # 5. 종목 개수가 많은 순서대로 정렬 (내림차순)
    summary_df = summary_df.sort_values(by='종목갯수', ascending=False)
    
    return summary_df

def apply_conditional_formatting(writer, sheet_name, df):
    """
    엑셀 시트에 조건부 서식(색상 강조)을 적용합니다.
    
    적용 규칙:
    1. '선정사유'가 'A'이면 '등락률' 셀을 노란색으로 칠함
    2. '선정사유'가 'B'이면 '거래량' 셀을 주홍색으로 칠함
    3. 테마 빈도수에 따라 테마 셀 색상 지정 (3개↑: 주홍, 2개: 노랑, 1개: 초록)
    
    Args:
        writer (pd.ExcelWriter): Pandas ExcelWriter 객체
        sheet_name (str): 서식을 적용할 시트 이름
        df (pd.DataFrame): 해당 시트에 기록된 원본 데이터프레임 (데이터 참조용)
    """
    # 워크시트 객체 가져오기
    ws = writer.sheets[sheet_name]
    
    # 색상 스타일 정의 (PatternFill 사용)
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # 노랑
    fill_orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid') # 주홍
    fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 초록
    
    # 헤더(첫 번째 행)에서 컬럼 이름과 인덱스 매핑 생성
    header_row = 1
    headers = {cell.value: i for i, cell in enumerate(ws[header_row], start=1)}
    
    # 주요 컬럼들의 엑셀 인덱스 찾기
    col_reason = headers.get('선정사유')
    col_fluctuation = headers.get('등락률')
    col_volume = headers.get('거래량')
    theme_cols_idx = [i for name, i in headers.items() if str(name).startswith('테마_')]
    
    # 테마 빈도수 계산 (전체 데이터 기준)
    # 1. 테마 컬럼 이름들 가져오기
    theme_col_names = [col for col in df.columns if col.startswith('테마_')]
    # 2. 2차원 데이터를 1차원 리스트로 펼치기 (flatten)
    all_themes = df[theme_col_names].values.flatten()
    # 3. 유효한 테마만 필터링
    all_themes = [t for t in all_themes if pd.notna(t) and t != '']
    # 4. 각 테마가 몇 번 등장했는지 카운트
    theme_counts = Counter(all_themes)
    
    # 데이터 행 순회 (헤더 다음인 2번째 행부터 끝까지)
    for row in range(2, ws.max_row + 1):
        # --- 규칙 1 & 2: 선정사유에 따른 하이라이트 ---
        if col_reason:
            reason_val = ws.cell(row=row, column=col_reason).value
            
            if reason_val == 'A' and col_fluctuation:
                # 선정사유 A -> 등락률 노란색
                ws.cell(row=row, column=col_fluctuation).fill = fill_yellow
            elif reason_val == 'B' and col_volume:
                # 선정사유 B -> 거래량 주홍색
                ws.cell(row=row, column=col_volume).fill = fill_orange
        
        # --- 규칙 3: 테마 빈도수에 따른 하이라이트 ---
        for col_idx in theme_cols_idx:
            cell = ws.cell(row=row, column=col_idx)
            theme_name = cell.value
            
            # 테마 이름이 있고, 카운트 정보가 있다면
            if theme_name and theme_name in theme_counts:
                count = theme_counts[theme_name]
                target_fill = None
                
                # 빈도수에 따라 색상 결정
                if count >= 3:
                    target_fill = fill_orange
                elif count == 2:
                    target_fill = fill_yellow
                elif count == 1:
                    target_fill = fill_green
                    
                # 결정된 색상이 있으면 적용
                if target_fill:
                    cell.fill = target_fill

def auto_adjust_column_width(writer):
    """
    엑셀 파일의 모든 시트에 대해 컬럼 너비를 내용에 맞춰 자동으로 조정합니다.
    한글 등 글자가 잘리지 않게 여유를 둡니다.
    
    Args:
        writer (pd.ExcelWriter): Pandas ExcelWriter 객체
    """
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        
        # 모든 열(column)을 순회하며 최대 길이 계산
        for column_cells in ws.columns:
            # 해당 열에서 가장 긴 글자수 찾기 (헤더 포함)
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            
            # 너비 계산 로직 보정
            # 기본 길이 * 1.5 + 2 (한글/특수문자 고려하여 넉넉하게)
            adjusted_width = (length * 1.5) + 2
            
            # 너무 넓어지지 않도록 최대 너비 제한 (예: 60)
            adjusted_width = min(adjusted_width, 60)
            
            # 해당 열의 너비 설정
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = adjusted_width
